import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from ..models import User
from .auth import get_current_user

router = APIRouter(prefix="/templates", tags=["templates"])

EMISSION_HEADERS = [
    "year", "sector", "reporting_boundary", "employee_count",
    "natural_gas_m3", "diesel_liters", "lpg_kg", "coal_tons", "company_vehicles_km",
    "electricity_kwh", "renewable_electricity_kwh",
    "business_flights_shorthaul", "business_flights_longhaul",
    "employee_commute_km", "waste_tons",
    "loan_portfolio_tl", "financed_emissions_co2e",
    "clinker_tons", "cement_tons",
    "electricity_generated_mwh", "renewable_capacity_mw",
]

HEADER_LABELS = {
    "year": "Yıl",
    "sector": "Sektör (banking/cement/energy/construction/retail/insurance/manufacturing/refinery)",
    "reporting_boundary": "Sınır (operational_control/financial_control/equity_share)",
    "employee_count": "Çalışan Sayısı",
    "natural_gas_m3": "Doğalgaz (m³/yıl)",
    "diesel_liters": "Dizel (litre/yıl)",
    "lpg_kg": "LPG (kg/yıl)",
    "coal_tons": "Kömür (ton/yıl)",
    "company_vehicles_km": "Şirket Araçları (km/yıl)",
    "electricity_kwh": "Satın Alınan Elektrik (kWh/yıl)",
    "renewable_electricity_kwh": "Yenilenebilir Elektrik - REC (kWh/yıl)",
    "business_flights_shorthaul": "İş Uçuşu Kısa Mesafe (km/yıl)",
    "business_flights_longhaul": "İş Uçuşu Uzun Mesafe (km/yıl)",
    "employee_commute_km": "Çalışan İşe Gidiş (km/yıl)",
    "waste_tons": "Atık (ton/yıl)",
    "loan_portfolio_tl": "Kredi Portföyü (milyon TL) — Bankacılık",
    "financed_emissions_co2e": "Finanse Edilmiş Emisyonlar (ton CO₂e) — Bankacılık",
    "clinker_tons": "Klinker Üretimi (ton/yıl) — Çimento",
    "cement_tons": "Çimento Üretimi (ton/yıl) — Çimento",
    "electricity_generated_mwh": "Elektrik Üretimi (MWh/yıl) — Enerji",
    "renewable_capacity_mw": "Yenilenebilir Kapasite (MW) — Enerji",
}

EXAMPLE_ROW = [
    2024, "manufacturing", "operational_control", 250,
    50000, 10000, 0, 0, 100000,
    500000, 0,
    50000, 20000,
    200000, 50,
    0, 0,
    0, 0,
    0, 0,
]


@router.get("/emissions")
async def download_emission_template(current_user: User = Depends(get_current_user)):  # noqa: ARG001
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emisyon Verisi"

    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    example_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    # Row 1: field keys (for parsing)
    for col, key in enumerate(EMISSION_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Row 2: human-readable labels
    label_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    for col, key in enumerate(EMISSION_HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=HEADER_LABELS.get(key, key))
        cell.fill = label_fill
        cell.font = Font(color="FFFFFF", size=9, italic=True)
        cell.alignment = Alignment(wrap_text=True)

    # Row 3: example data
    for col, val in enumerate(EXAMPLE_ROW, start=1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.fill = example_fill

    # Column widths
    for col in range(1, len(EMISSION_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sustain-emisyon-sablonu.xlsx"},
    )
